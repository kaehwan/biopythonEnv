#! Python 3.7, openpyxl 3.0.3
# - Copy and Paste Ranges using OpenPyXl library

import sys
import openpyxl


def main():
    # Prepare the spreadsheets to copy from and paste too.
    userParameters = sys.argv[1:]

    try:
        if len(userParameters) != 4:
            print("You must have 4 inputs:",
                  "(1) Source Excel file (full directory); (2) Source Excel sheet name;",
                  "(3) Destination Excel file (full directory); (4) Destination Excel sheet name")
        else:
            print("Inputs valid...")

            # File to be copied
            wb = openpyxl.load_workbook(userParameters[0])  # Add file name
            sheet = wb[userParameters[1]]  # Add Sheet name

            # File to be pasted into
            template = openpyxl.load_workbook(userParameters[2])  # Add file name
            temp_sheet = template[userParameters[3]]  # Add Sheet name

            # Copy range of cells as a nested list
            # Takes: start cell, end cell, and sheet you want to copy from.
            def copyRange(startCol, startRow, endCol, endRow, sheet):
                rangeSelected = []
                # Loops through selected Rows
                for i in range(startRow, endRow + 1, 1):
                    # Appends the row to a RowSelected list
                    rowSelected = []
                    for j in range(startCol, endCol + 1, 1):
                        rowSelected.append(sheet.cell(row=i, column=j).value)
                    # Adds the RowSelected List and nests inside the rangeSelected
                    rangeSelected.append(rowSelected)

                return rangeSelected

            # Paste range
            # Paste data from copyRange into template sheet
            def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
                countRow = 0
                for i in range(startRow, endRow + 1, 1):
                    countCol = 0
                    for j in range(startCol, endCol + 1, 1):
                        sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
                        countCol += 1
                    countRow += 1

            def createData():
                print("Processing...")
                selectedRange = copyRange(1, 2, 4, 14, sheet)  # Change the 4 number values
                pasteRange(1, 3, 4, 15, temp_sheet, selectedRange)
                # You can save the template as another file to create a new file here too.s
                template.save(userParameters[2])
                print("Range copied and pasted!")

            createData()
    except:
        print("There is a problem!")


if __name__ == "__main__":
    main()
